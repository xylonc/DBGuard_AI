"""
Evidence & Reporting Service
Handles evidence collection, screenshot embedding, hash verification,
Excel checklist generation, and final audit package assembly.
"""
from __future__ import annotations

import hashlib
import io
import json
import logging
import os
import shutil
import subprocess
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, SecretStr

logger = logging.getLogger("dbguard.evidence")


# ─── Enums ───────────────────────────────────────────────────────────

class EvidenceType(str, Enum):
    SCREENSHOT = "screenshot"
    QUERY_OUTPUT = "query_output"
    COMMAND_OUTPUT = "command_output"
    CONFIGURATION_EXCERPT = "configuration_excerpt"
    POLICY_DOCUMENT = "policy_document"
    CHANGE_TICKET = "change_ticket"
    REVIEWER_ATTESTATION = "reviewer_attestation"
    OTHER = "other"


class EvidenceStatus(str, Enum):
    PENDING = "PENDING"
    ACCEPTED = "ACCEPTED"
    REJECTED = "REJECTED"
    INSUFFICIENT = "INSUFFICIENT"
    REDACTION_REQUIRED = "REDACTION_REQUIRED"
    RECAPTURE_REQUIRED = "RECAPTURE_REQUIRED"


class ReviewerDecision(str, Enum):
    ACCEPT = "ACCEPT"
    REJECT = "REJECT"
    INSUFFICIENT = "INSUFFICIENT"
    REDACTION_REQUIRED = "REDACTION_REQUIRED"
    RECAPTURE_REQUIRED = "RECAPTURE_REQUIRED"


class TargetVerificationStatus(str, Enum):
    TWIN_VALIDATED = "TWIN_VALIDATED"
    TARGET_AUTOMATICALLY_VERIFIED = "TARGET_AUTOMATICALLY_VERIFIED"
    TARGET_MANUALLY_VERIFIED = "TARGET_MANUALLY_VERIFIED"
    TARGET_VERIFICATION_FAILED = "TARGET_VERIFICATION_FAILED"
    TARGET_EVIDENCE_MISSING = "TARGET_EVIDENCE_MISSING"


# ─── Data Models ─────────────────────────────────────────────────────

class EvidenceItem(BaseModel):
    """Individual evidence record."""
    evidence_id: str
    run_id: str
    control_id: str
    evidence_type: EvidenceType
    description: str
    original_filename: Optional[str] = None
    file_hash: Optional[str] = None  # SHA-256
    capture_timestamp: datetime = Field(default_factory=datetime.utcnow)
    operator: str
    redaction_status: str = "not_checked"  # not_checked, redacted, verified
    reviewer: Optional[str] = None
    review_status: EvidenceStatus = EvidenceStatus.PENDING
    review_comments: Optional[str] = None
    review_timestamp: Optional[datetime] = None
    checklist_row: Optional[int] = None
    change_ticket_reference: Optional[str] = None


class ControlChecklistRow(BaseModel):
    """A row in the Excel evidence checklist for a control."""
    sequence: int
    control_id: str
    cis_cat_item_id: Optional[str] = None
    control_title: str
    is_manual: bool = False
    applicability: str = "applicable"
    baseline_target_status: str = "not_assessed"
    twin_baseline_status: str = "not_assessed"
    twin_final_status: str = "not_assessed"
    proposed_remediation: str = ""
    impact_rating: str = "low"
    dependencies: str = ""
    human_approval_required: bool = False
    approval_status: str = "pending"
    apply_script_ref: str = ""
    verification_script_ref: str = ""
    rollback_script_ref: str = ""
    reload_required: bool = False
    restart_required: bool = False
    target_implementation_status: str = "not_started"  # not_started, started, applied, verified, failed, rolled_back
    target_implementation_timestamp: Optional[datetime] = None
    target_verification_result: str = ""
    post_implementation_cis_cat_status: str = "not_assessed"
    evidence_required: str = ""
    evidence_id: Optional[str] = None
    exception_status: str = "none"  # none, proposed, accepted, rejected
    exception_justification: Optional[str] = None
    operator: Optional[str] = None
    reviewer: Optional[str] = None
    reviewer_decision: Optional[str] = None
    reviewer_decision_timestamp: Optional[datetime] = None
    notes: str = ""


class EvidenceRecord(BaseModel):
    """Immutable record of evidence with hash and metadata."""
    evidence_id: str
    run_id: str
    file_path: str
    original_filename: str
    file_size_bytes: int
    file_hash: str  # SHA-256
    content_type: str
    capture_timestamp: datetime
    redaction_status: str = "not_checked"
    malware_scan_status: str = "pending"  # pending, passed, failed
    checklist_association: Optional[str] = None
    reviewer_status: str = "pending"


class RunSummary(BaseModel):
    """Summary of an assessment run for the Excel workbook."""
    assessment_run_id: str
    target_identifier: str
    postgresql_version: str
    postgresql_distribution: str = "community"
    image_catalog_profile: str = ""
    cis_benchmark: str = "CIS PostgreSQL Benchmark"
    cis_version: str = "1.0.0"
    snapshot_date: Optional[datetime] = None
    twin_assessment_date: Optional[datetime] = None
    target_implementation_date: Optional[datetime] = None
    post_implementation_assessment_date: Optional[datetime] = None
    operator: str = ""
    reviewer: str = ""
    baseline_target_score: float = 0.0
    twin_post_hardening_score: float = 0.0
    final_target_score: Optional[float] = None
    assessment_coverage: float = 0.0
    twin_fidelity: str = "unknown"
    package_hash: str = ""


# ─── Evidence Service ────────────────────────────────────────────────

class EvidenceService:
    """
    Handles evidence collection, verification, Excel checklist generation,
    and final audit package assembly.
    
    Key principles:
    - Original evidence files are preserved (never modified)
    - SHA-256 hash is calculated for every uploaded file
    - Screenshots are embedded in Excel with resized previews
    - Manual evidence requires human review before control can be marked verified
    - Missing evidence remains visibly incomplete in all reports
    """
    
    def __init__(self, evidence_dir: str = "evidence"):
        self.evidence_dir = evidence_dir
        self.checklist_dir = "reports"
        os.makedirs(self.evidence_dir, exist_ok=True)
        os.makedirs(self.checklist_dir, exist_ok=True)
    
    def save_evidence_file(self, evidence_id: str, file_data: bytes, 
                           original_filename: str, evidence_type: EvidenceType,
                           run_id: str, control_id: str, operator: str) -> EvidenceRecord:
        """
        Save an evidence file and calculate its SHA-256 hash.
        
        Returns an EvidenceRecord with hash and metadata.
        """
        # Calculate hash before saving
        file_hash = hashlib.sha256(file_data).hexdigest()
        
        # Determine content type
        content_type = self._detect_content_type(file_data, original_filename)
        
        # Save file with evidence ID prefix
        safe_filename = self._sanitize_filename(original_filename)
        save_path = os.path.join(
            self.evidence_dir,
            f"{evidence_id}-{safe_filename}"
        )
        
        with open(save_path, "wb") as f:
            f.write(file_data)
        
        file_size = len(file_data)
        
        record = EvidenceRecord(
            evidence_id=evidence_id,
            run_id=run_id,
            file_path=save_path,
            original_filename=original_filename,
            file_size_bytes=file_size,
            file_hash=file_hash,
            content_type=content_type,
            capture_timestamp=datetime.utcnow(),
        )
        
        logger.info(f"Evidence saved: {evidence_id} ({safe_filename}, {file_hash[:16]}..., {file_size} bytes)")
        return record
    
    def upload_screenshot(self, screenshot_data: bytes, original_filename: str,
                          control_id: str, run_id: str, operator: str) -> EvidenceRecord:
        """
        Upload a screenshot for a manual control.
        
        Preserves original, calculates hash, embeds preview.
        An uploaded screenshot is NOT automatically sufficient evidence —
        it requires reviewer acceptance.
        """
        # Validate file type
        if not self._is_valid_image(screenshot_data):
            raise ValueError("Uploaded file is not a valid image")
        
        # Generate evidence ID
        evidence_id = f"EVID-{run_id}-{control_id}-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
        
        # Save and hash
        record = self.save_evidence_file(
            evidence_id=evidence_id,
            file_data=screenshot_data,
            original_filename=original_filename,
            evidence_type=EvidenceType.SCREENSHOT,
            run_id=run_id,
            control_id=control_id,
            operator=operator,
        )
        
        logger.info(f"Screenshot uploaded: {evidence_id}")
        return record
    
    def verify_evidence_hash(self, evidence_id: str, expected_hash: str) -> bool:
        """Verify that an evidence file's hash matches the expected value."""
        evidence_path = os.path.join(self.evidence_dir, f"{evidence_id}-*")
        
        # Find the evidence file
        import glob
        files = glob.glob(os.path.join(self.evidence_dir, f"{evidence_id}-*"))
        
        if not files:
            logger.error(f"Evidence file not found: {evidence_id}")
            return False
        
        # Calculate hash of existing file
        with open(files[0], "rb") as f:
            actual_hash = hashlib.sha256(f.read()).hexdigest()
        
        if actual_hash == expected_hash:
            logger.info(f"Hash verified: {evidence_id}")
            return True
        else:
            logger.error(f"Hash mismatch for {evidence_id}: expected {expected_hash[:16]}, got {actual_hash[:16]}")
            return False
    
    def record_reviewer_decision(self, evidence_id: str, decision: ReviewerDecision,
                                  reviewer: str, comments: Optional[str] = None) -> Tuple[bool, str]:
        """
        Record a reviewer's decision on evidence.
        
        Decisions:
        - ACCEPT: Evidence is sufficient, control can be marked verified
        - REJECT: Evidence is invalid or tampered
        - INSUFFICIENT: Evidence exists but doesn't fully verify the control
        - REDACTION_REQUIRED: Evidence contains sensitive data that must be redacted
        - RECAPTURE_REQUIRED: Evidence is unclear or incomplete
        """
        # Find and update evidence record
        import glob
        files = glob.glob(os.path.join(self.evidence_dir, f"{evidence_id}-*"))
        
        if not files:
            return False, f"Evidence not found: {evidence_id}"
        
        # Update record status
        # In production, this would write to a database or manifest file
        manifest_path = os.path.join(self.evidence_dir, "manifest.json")
        
        manifest = {}
        if os.path.exists(manifest_path):
            with open(manifest_path, "r") as f:
                manifest = json.load(f)
        
        manifest[evidence_id] = {
            "status": decision.value,
            "reviewer": reviewer,
            "review_comments": comments,
            "review_timestamp": datetime.utcnow().isoformat(),
        }
        
        with open(manifest_path, "w") as f:
            json.dump(manifest, f, indent=2)
        
        logger.info(f"Reviewer decision recorded: {evidence_id} -> {decision.value}")
        return True, f"Decision recorded: {decision.value}"
    
    def generate_excel_checklist(self, run_summary: RunSummary, 
                                  checklist_rows: List[ControlChecklistRow],
                                  evidence_records: List[EvidenceRecord]) -> Tuple[bool, str]:
        """
        Generate an Excel evidence checklist with embedded screenshot previews.
        
        Worksheets:
        1. Run Summary
        2. Control Checklist
        3. Manual Evidence
        4. Exceptions
        5. Evidence Index
        6. Change History
        
        Screenshots are resized for embedding, original files preserved separately.
        """
        output_path = os.path.join(
            self.checklist_dir,
            f"DBGuard-Evidence-Checklist-{run_summary.assessment_run_id}.xlsx"
        )
        
        try:
            wb = Workbook()
            
            # ─── Worksheet 1: Run Summary ───
            ws_summary = wb.active
            ws_summary.title = "Run Summary"
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 60
            
            summary_data = [
                ("Assessment Run ID", run_summary.assessment_run_id),
                ("Target Identifier", run_summary.target_identifier),
                ("PostgreSQL Version", run_summary.postgresql_version),
                ("PostgreSQL Distribution", run_summary.postgresql_distribution),
                ("Image Catalog Profile", run_summary.image_catalog_profile),
                ("CIS Benchmark", run_summary.cis_benchmark),
                ("CIS Version", run_summary.cis_version),
                ("Snapshot Date", run_summary.snapshot_date.isoformat() if run_summary.snapshot_date else ""),
                ("Twin Assessment Date", run_summary.twin_assessment_date.isoformat() if run_summary.twin_assessment_date else ""),
                ("Target Implementation Date", run_summary.target_implementation_date.isoformat() if run_summary.target_implementation_date else ""),
                ("Post-Implementation Date", run_summary.post_implementation_assessment_date.isoformat() if run_summary.post_implementation_assessment_date else ""),
                ("Operator", run_summary.operator),
                ("Reviewer", run_summary.reviewer),
                ("Baseline Target Score", f"{run_summary.baseline_target_score}%"),
                ("Twin Post-Hardening Score", f"{run_summary.twin_post_hardening_score}%"),
                ("Final Target Score", f"{run_summary.final_target_score}%" if run_summary.final_target_score else "Pending"),
                ("Assessment Coverage", f"{run_summary.assessment_coverage}%"),
                ("Twin Fidelity", run_summary.twin_fidelity),
                ("Package Hash", run_summary.package_hash),
            ]
            
            ws_summary.append(["Field", "Value"])
            for field, value in summary_data:
                ws_summary.append([field, value])
            
            # ─── Worksheet 2: Control Checklist ───
            ws_checklist = wb.create_sheet("Control Checklist")
            
            # Column headers
            headers = [
                "Sequence", "Control ID", "CIS-CAT Item", "Control Title",
                "Automated/Manual", "Applicability", "Baseline Target", "Twin Baseline",
                "Twin Final", "Proposed Remediation", "Impact Rating", "Dependencies",
                "Human Approval", "Approval Status", "Apply Script", "Verification Script",
                "Rollback Script", "Reload Required", "Restart Required",
                "Target Implementation Status", "Target Timestamp",
                "Target Verification Result", "Post-Impl CIS-CAT",
                "Evidence Required", "Evidence ID", "Exception Status",
                "Exception Justification", "Operator", "Reviewer",
                "Reviewer Decision", "Review Timestamp", "Notes"
            ]
            
            ws_checklist.append(headers)
            for col_letter in get_column_letter(ws_checklist.max_column):
                ws_checklist.column_dimensions[col_letter].width = 18
            
            # Column widths (override for wider columns)
            ws_checklist.column_dimensions['H'].width = 30  # Proposed Remediation
            ws_checklist.column_dimensions['AG'].width = 30  # Notes
            
            # Checklist rows
            for row in checklist_rows:
                checklist_row = [
                    row.sequence, row.control_id, row.cis_cat_item_id,
                    row.control_title, "Manual" if row.is_manual else "Automated",
                    row.applicability, row.baseline_target_status,
                    row.twin_baseline_status, row.twin_final_status,
                    row.proposed_remediation, row.impact_rating,
                    row.dependencies, "Yes" if row.human_approval_required else "No",
                    row.approval_status, row.apply_script_ref,
                    row.verification_script_ref, row.rollback_script_ref,
                    "Yes" if row.reload_required else "No",
                    "Yes" if row.restart_required else "No",
                    row.target_implementation_status,
                    row.target_implementation_timestamp.isoformat() if row.target_implementation_timestamp else "",
                    row.target_verification_result, row.post_implementation_cis_cat_status,
                    row.evidence_required, row.evidence_id,
                    row.exception_status, row.exception_justification or "",
                    row.operator, row.reviewer,
                    row.reviewer_decision,
                    row.reviewer_decision_timestamp.isoformat() if row.reviewer_decision_timestamp else "",
                    row.notes,
                ]
                ws_checklist.append(checklist_row)
            
            # ─── Worksheet 3: Manual Evidence ───
            ws_evidence = wb.create_sheet("Manual Evidence")
            evidence_headers = [
                "Evidence ID", "Control ID", "Description", "Evidence Type",
                "Capture Instructions", "Screenshot Preview", "Original Path",
                "Evidence Hash", "Capture Timestamp", "Operator",
                "Redaction Status", "Reviewer", "Evidence Status", "Notes"
            ]
            ws_evidence.append(evidence_headers)
            
            # Add manual evidence records with embedded previews
            for idx, record in enumerate(evidence_records):
                if record.content_type.startswith("image/"):
                    evidence_row = [
                        record.evidence_id, "", record.original_filename,
                        "Screenshot", "", "", record.file_path,
                        record.file_hash, record.capture_timestamp.isoformat(),
                        "", record.redaction_status, "", record.malware_scan_status, ""
                    ]
                    ws_evidence.append(evidence_row)
                    
                    # Embed screenshot preview (first 20 chars of filename as path hint)
                    try:
                        preview_path = os.path.join(self.evidence_dir, f"{record.evidence_id}-*")
                        import glob as g
                        matches = g.glob(preview_path)
                        if matches:
                            img = XLImage(matches[0])
                            img.width = 200
                            img.height = 150
                            ws_evidence.add_image(img, f"F{idx + 2}")
                    except Exception as e:
                        logger.warning(f"Could not embed screenshot preview: {e}")
        
                else:
                    evidence_row = [
                        record.evidence_id, "", record.original_filename,
                        record.content_type, "", "", record.file_path,
                        record.file_hash, record.capture_timestamp.isoformat(),
                        "", record.redaction_status, "", record.malware_scan_status, ""
                    ]
                    ws_evidence.append(evidence_row)
            
            # ─── Worksheet 4: Exceptions ───
            ws_exceptions = wb.create_sheet("Exceptions")
            ws_exceptions.append(["Control ID", "Reason", "Attempts Made", "Residual Risk",
                                  "Compensating Controls", "Business Justification",
                                  "Owner", "Expiry Date", "Approval Status", "Supporting Evidence"])
            # (Exception rows would be populated from the assessment data)
            
            # ─── Worksheet 5: Evidence Index ───
            ws_index = wb.create_sheet("Evidence Index")
            ws_index.append(["Evidence ID", "File", "Hash", "Type", "Status", "Reviewer"])
            for record in evidence_records:
                ws_index.append([
                    record.evidence_id, record.original_filename,
                    record.file_hash, record.content_type,
                    record.malware_scan_status, ""
                ])
            
            # ─── Worksheet 6: Change History ───
            ws_history = wb.create_sheet("Change History")
            ws_history.append(["Timestamp", "Action", "Control ID", "By", "Status", "Notes"])
            # (Change history rows would be populated from the assessment data)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel checklist generated: {output_path}")
            return True, output_path
            
        except Exception as e:
            logger.error(f"Failed to generate Excel checklist: {e}")
            return False, str(e)
    
    def generate_cis_cat_import_report(self, cis_cat_report_path: str,
                                        run_id: str) -> Tuple[bool, Dict[str, Any]]:
        """
        Import and parse a CIS-CAT report.
        
        Returns:
            (success, parsed_data)
        
        Parses:
        - Benchmark and version
        - Target identity
        - Assessment timestamp
        - Imported controls
        - Unmapped controls
        - Parsing errors
        """
        parsed_data = {
            "benchmark": "",
            "version": "",
            "target_identity": "",
            "assessment_timestamp": "",
            "imported_controls": 0,
            "unmapped_controls": 0,
            "parsing_errors": [],
            "score": 0.0,
        }
        
        try:
            # CIS-CAT reports are typically JSON or XML
            if cis_cat_report_path.endswith(".json"):
                with open(cis_cat_report_path, "r") as f:
                    data = json.load(f)
                
                # Parse JSON report
                parsed_data["benchmark"] = data.get("benchmark", "CIS PostgreSQL Benchmark")
                parsed_data["version"] = data.get("version", "1.0.0")
                parsed_data["target_identity"] = data.get("target_host", "unknown")
                parsed_data["assessment_timestamp"] = data.get("timestamp", datetime.utcnow().isoformat())
                parsed_data["score"] = data.get("score", 0.0)
                
                controls = data.get("results", {}).get("controls", [])
                parsed_data["imported_controls"] = len(controls)
                
                for control in controls:
                    if control.get("status") == "notapplicable":
                        pass  # Count as imported but not applicable
                    elif control.get("status") == "fail":
                        pass  # Count as imported but failed
            
            elif cis_cat_report_path.endswith(".xml"):
                # Parse XML report
                import xml.etree.ElementTree as ET
                tree = ET.parse(cis_cat_report_path)
                root = tree.getroot()
                
                # Extract benchmark info
                benchmark_elem = root.find(".//benchmark")
                if benchmark_elem is not None:
                    parsed_data["benchmark"] = benchmark_elem.get("id", "unknown")
                    parsed_data["version"] = benchmark_elem.get("version", "1.0.0")
                
                # Extract controls
                controls = root.findall(".//result")
                parsed_data["imported_controls"] = len(controls)
                
                # Count failed controls
                failed = [c for c in controls if c.get("outcome") == "fail"]
                parsed_data["failed_controls"] = len(failed)
            
            else:
                parsed_data["parsing_errors"].append(f"Unsupported format: {cis_cat_report_path}")
                return False, parsed_data
            
            logger.info(f"CIS-CAT report imported: {parsed_data['imported_controls']} controls parsed")
            return True, parsed_data
            
        except Exception as e:
            logger.error(f"Failed to import CIS-CAT report: {e}")
            parsed_data["parsing_errors"].append(str(e))
            return False, parsed_data
    
    def generate_final_report(self, run_id: str, run_summary: RunSummary,
                               checklist_rows: List[ControlChecklistRow],
                               evidence_records: List[EvidenceRecord]) -> str:
        """
        Generate the final audit report.
        
        Generates both HTML and PDF (via LibreOffice conversion).
        """
        report_dir = os.path.join(self.checklist_dir, f"DBGuard-Assessment-{run_id}")
        os.makedirs(report_dir, exist_ok=True)
        
        # Generate HTML report
        html_path = os.path.join(report_dir, "assessment-report.html")
        self._generate_html_report(html_path, run_summary, checklist_rows, evidence_records)
        
        # Generate PDF (if LibreOffice is available)
        pdf_path = os.path.join(report_dir, "assessment-report.pdf")
        try:
            subprocess.run([
                "libreoffice", "--headless", "--convert-to", "pdf",
                html_path, "--outdir", report_dir
            ], capture_output=True, text=True, timeout=60)
            if os.path.exists(pdf_path):
                logger.info(f"PDF report generated: {pdf_path}")
        except Exception as e:
            logger.warning(f"PDF generation failed (HTML available): {e}")
        
        logger.info(f"Final report generated: {report_dir}")
        return report_dir
    
    def _generate_html_report(self, path: str, summary: RunSummary,
                                rows: List[ControlChecklistRow],
                                evidence: List[EvidenceRecord]) -> None:
        """Generate an HTML report."""
        html = f"""<!DOCTYPE html>
<html>
<head>
    <title>DBGuardAI Assessment Report — {summary.assessment_run_id}</title>
    <style>
        body {{ font-family: Inter, Arial, sans-serif; margin: 40px; color: #333; }}
        h1 {{ color: #1a365d; border-bottom: 2px solid #319795; padding-bottom: 10px; }}
        h2 {{ color: #2d3748; margin-top: 30px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #e2e8f0; padding: 8px 12px; text-align: left; }}
        th {{ background-color: #f7fafc; font-weight: 600; }}
        .score {{ font-size: 2em; font-weight: bold; }}
        .score-green {{ color: #38a169; }}
        .score-amber {{ color: #d69e2e; }}
        .score-red {{ color: #e53e3e; }}
        .status-passed {{ color: #38a169; font-weight: bold; }}
        .status-failed {{ color: #e53e3e; font-weight: bold; }}
        .status-manual {{ color: #d69e2e; font-weight: bold; }}
        .status-pending {{ color: #718096; font-weight: bold; }}
        .disclaimer {{ background: #fff5f5; border-left: 4px solid #e53e3e; padding: 15px; margin: 20px 0; }}
    </style>
</head>
<body>
    <h1>DBGuardAI Assessment Report</h1>
    <p><strong>Run ID:</strong> {summary.assessment_run_id}</p>
    <p><strong>Target:</strong> {summary.target_identifier}</p>
    <p><strong>PostgreSQL:</strong> {summary.postgresql_version}</p>
    
    <h2>Summary</h2>
    <div style="display: flex; gap: 20px; margin: 20px 0;">
        <div style="background: #f7fafc; padding: 20px; border-radius: 8px;">
            <p class="score score-amber">{summary.baseline_target_score}%</p>
            <p>Baseline Target Score</p>
        </div>
        <div style="background: #f7fafc; padding: 20px; border-radius: 8px;">
            <p class="score score-green">{summary.twin_post_hardening_score}%</p>
            <p>Twin Post-Hardening Score</p>
        </div>
        <div style="background: #f7fafc; padding: 20px; border-radius: 8px;">
            <p class="score score-amber">{summary.assessment_coverage}%</p>
            <p>Assessment Coverage</p>
        </div>
    </div>
    
    <h2>Controls</h2>
    <table>
        <tr>
            <th>Sequence</th><th>Control ID</th><th>Title</th>
            <th>Twin Status</th><th>Target Status</th>
            <th>Evidence</th><th>Exception</th>
        </tr>"""
        
        for row in rows:
            twin_class = "passed" if "passed" in row.twin_final_status.lower() else "failed"
            target_class = "pending" if "not_started" in row.target_implementation_status.lower() else "pending"
            
            html += f"""
        <tr>
            <td>{row.sequence}</td>
            <td>{row.control_id}</td>
            <td>{row.control_title}</td>
            <td class="status-{twin_class}">{row.twin_final_status}</td>
            <td class="status-{target_class}">{row.target_implementation_status}</td>
            <td>{"✓" if row.evidence_id else "—"}</td>
            <td>{"Exception" if row.exception_status != "none" else "—"}</td>
        </tr>"""
        
        html += """
    </table>
    
    <h2>Warnings</h2>
    <div class="disclaimer">
        <strong>Important:</strong> DBGuard does not modify production databases directly.
        Approved changes require human review and must be applied through your organisation's
        normal change-management procedures. Twin testing results do not guarantee production safety.
    </div>
    
    <footer style="margin-top: 40px; color: #718096; font-size: 0.9em;">
        <p>Generated by DBGuardAI — Database Security Hardening & Decision Support Platform</p>
        <p>Report hash will be calculated when package is finalized.</p>
    </footer>
</body>
</html>"""
        
        with open(path, "w") as f:
            f.write(html)
    
    def _detect_content_type(self, file_data: bytes, filename: str) -> str:
        """Detect content type from file data or extension."""
        if filename.lower().endswith(".png"):
            return "image/png"
        elif filename.lower().endswith((".jpg", ".jpeg")):
            return "image/jpeg"
        elif filename.lower().endswith(".pdf"):
            return "application/pdf"
        elif filename.lower().endswith(".txt"):
            return "text/plain"
        elif file_data.startswith(b"%PDF"):
            return "application/pdf"
        elif file_data.startswith(b"\x89PNG"):
            return "image/png"
        elif file_data.startswith(b"\xff\xd8\xff"):
            return "image/jpeg"
        return "application/octet-stream"
    
    def _is_valid_image(self, file_data: bytes) -> bool:
        """Check if file data is a valid image."""
        return (file_data.startswith(b"\x89PNG") or
                file_data.startswith(b"\xff\xd8\xff") or
                file_data.startswith(b"RIFF"))
    
    def _sanitize_filename(self, filename: str) -> str:
        """Sanitize filename for safe storage."""
        # Remove path separators and special characters
        safe = filename.replace("/", "_").replace("\\", "_").replace("..", "_")
        # Limit length
        if len(safe) > 200:
            safe = safe[:200]
        return safe
